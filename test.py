import streamlit as st
import pandas as pd
import os
import re
import base64
import requests
from datetime import datetime
from openpyxl import load_workbook

# --- Page Config ---
st.set_page_config(page_title="🚛 Mahindra Docket Audit Tool - CV", layout="centered" )

# --- Constants ---
DATA_DIR = "Data/Discount_Cheker"
FILE_PATTERN = r"CV Discount Check Master File (\d{2})\.(\d{2})\.(\d{4})\.xlsx"
SHEET_NAME = "Sheet1"
HEADER_ROW = 1

# --- Global Styling ---
st.markdown("""
<style>
:root {
    --title-size: 40px;
    --subtitle-size: 20px;
    --caption-size: 16px;
    --label-size: 14px;
    --select-font-size: 15px;
    --table-font-size: 14px;
    --variant-title-size: 24px;
}
.block-container { padding-top: 0rem; }
header {visibility: hidden;}
h1 { font-size: var(--title-size) !important; }
h2 { font-size: var(--subtitle-size) !important; }
h3 { font-size: var(--variant-title-size) !important; }
.stCaption { font-size: var(--caption-size) !important; }
.stSelectbox label {
    font-size: var(--label-size) !important;
    font-weight: 600 !important;
}

/* ✅ MINIMAL DROPDOWN STYLING */
.stSelectbox div[data-baseweb="select"] > div {
    font-size: 15px !important;
    font-weight: bold !important;
    padding-top: 2px !important;
    padding-bottom: 2px !important;
    line-height: 1 !important;
    min-height: 24px !important;
}

/* Adaptive color based on theme */
[data-theme="light"] .stSelectbox div[data-baseweb="select"] > div {
    color: black !important;
    background-color: #f3f4f6 !important;
}
[data-theme="dark"] .stSelectbox div[data-baseweb="select"] > div {
    color: white !important;
    background-color: #333 !important;
}
.stSelectbox div[data-baseweb="select"] {
    align-items: center !important;
    height: 28px !important;
}

.stSelectbox [data-baseweb="option"]:hover {
    background-color: #f0f0f0 !important;
    font-weight: 600 !important;
}
/* Light mode styling */
[data-theme="light"] .stSelectbox div[data-baseweb="select"] > div {
    color: black !important;
    background-color: #f3f4f6 !important; /* Light background */
    font-weight: bold !important;
}

/* Dark mode styling */
[data-theme="dark"] .stSelectbox div[data-baseweb="select"] > div {
    color: white !important;
    background-color: #333 !important; /* Dark background */
    font-weight: bold !important;
}

/* Hover styling for options (common to both) */
.stSelectbox [data-baseweb="option"]:hover {
    background-color: #e0e0e0 !important;
    font-weight: 600 !important;
}

/* Important Points Table  */
/* ----------------------- */
.iptable { border-collapse: collapse; width: 100%; font-weight: bold; font-size: 14px; }
.iptable th { background-color: #e65100; color: white; padding: 4px 6px; text-align: left; }
.iptable td { background-color: #fff3e0; padding: 4px 6px; text-align: left; color: black; }
.iptable, .iptable th, .iptable td { border: 1px solid #000; }

</style>
""", unsafe_allow_html=True)


# --- Admin Authentication ---
def check_admin_password():
    correct_password = st.secrets["auth"]["admin_password"]
    if "admin_authenticated" not in st.session_state:
        st.session_state["admin_authenticated"] = False

    if not st.session_state["admin_authenticated"]:
        with st.sidebar.expander("🔐 Admin Login", expanded=False):
            pwd = st.text_input("Enter admin password:", type="password", key="admin_pwd")
            if st.button("Login", key="admin_login_btn"):
                if pwd == correct_password:
                    st.session_state["admin_authenticated"] = True
                    st.rerun()
                else:
                    st.error("❌ Incorrect password.")
        return False
    return True

def logout_admin():
    if st.session_state.get("admin_authenticated", False):
        if st.sidebar.button("🔓 Logout Admin"):
            st.session_state["admin_authenticated"] = False
            st.rerun()

# --- GitHub Upload + Cleanup ---
def upload_to_github(file_path, filename):
    try:
        token = st.secrets["github"]["token"]
        username = st.secrets["github"]["username"]
        repo = st.secrets["github"]["repo"]
        branch = st.secrets["github"].get("branch", "main")
        github_dir = "Data/Discount_Cheker"

        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github+json"
        }

        github_file_path = f"{github_dir}/{filename}"
        with open(file_path, "rb") as f:
            content = base64.b64encode(f.read()).decode()

        url = f"https://api.github.com/repos/{username}/{repo}/contents/{github_file_path}"
        check = requests.get(url, headers=headers)
        sha = check.json().get("sha") if check.status_code == 200 else None

        payload = {
            "message": f"Upload Excel file {filename}",
            "content": content,
            "branch": branch
        }
        if sha:
            payload["sha"] = sha

        put = requests.put(url, headers=headers, json=payload)
        if put.status_code not in [200, 201]:
            st.sidebar.error(f"❌ GitHub upload failed: {put.json().get('message')}")
            return

        st.sidebar.success(f"✅ Uploaded to GitHub: {filename}")

        # Cleanup old files
        list_url = f"https://api.github.com/repos/{username}/{repo}/contents/{github_dir}"
        files_resp = requests.get(list_url, headers=headers)
        if files_resp.status_code != 200:
            st.sidebar.warning("⚠️ Could not fetch file list from GitHub.")
            return

        files_data = files_resp.json()
        excel_files = []
        for item in files_data:
            fname = item["name"]
            match = re.match(FILE_PATTERN, fname)
            if match:
                try:
                    day, month, year = match.groups()
                    fdate = datetime.strptime(f"{day}.{month}.{year}", "%d.%m.%Y")
                    excel_files.append((fname, fdate, item["sha"]))
                except:
                    continue

        excel_files.sort(key=lambda x: x[1], reverse=True)
        for fname, _, sha_to_delete in excel_files[5:]:
            del_url = f"https://api.github.com/repos/{username}/{repo}/contents/{github_dir}/{fname}"
            del_payload = {
                "message": f"Auto-delete old Excel file: {fname}",
                "sha": sha_to_delete,
                "branch": branch
            }
            requests.delete(del_url, headers=headers, json=del_payload)

    except Exception as e:
        st.sidebar.error(f"❌ GitHub Error: {str(e)}")

# --- Upload Section (Admin Only) ---
if check_admin_password():
    st.sidebar.header("📂 File Upload (Admin Only)")
    uploaded_file = st.sidebar.file_uploader("Upload New Excel File", type=["xlsx"])
    if uploaded_file:
        os.makedirs(DATA_DIR, exist_ok=True)
        save_path = os.path.join(DATA_DIR, uploaded_file.name)
        with open(save_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        upload_to_github(save_path, uploaded_file.name)
        st.rerun()
logout_admin()


# --- Government Services (Sidebar Shortcuts) ---
st.sidebar.markdown("---")
st.sidebar.markdown("### 🗂️ Government Services")

with st.sidebar:
    st.link_button("🏦 BLP Gujarat - Application Status", "https://blp.gujarat.gov.in/appstatussearch.php")
    st.link_button("🏢 Udyam Registration Verification", "https://udyamregistration.gov.in/Government-India/Ministry-MSME-registration.htm")
    st.link_button("🧾 Aadhaar–PAN Link Status", "https://eportal.incometax.gov.in/iec/foservices/#/pre-login/link-aadhaar-status")

# --- File Listing ---
def extract_date_from_filename(filename):
    match = re.search(FILE_PATTERN, filename)
    if match:
        try:
            day, month, year = match.groups()
            return datetime.strptime(f"{day}.{month}.{year}", "%d.%m.%Y")
        except:
            return None
    return None

files = [(f, extract_date_from_filename(f)) for f in os.listdir(DATA_DIR) if re.match(FILE_PATTERN, f)]
files = sorted([f for f in files if f[1]], key=lambda x: x[1], reverse=True)[:5]
if not files:
    st.error("❌ No valid Excel files found.")
    st.stop()

file_labels = [f"{fname} ({dt.strftime('%d-%b-%Y')})" for fname, dt in files]
file_map = {label: fname for label, (fname, _) in zip(file_labels, files)}

# --- Title ---
st.markdown(
    "<h1>🚛 Mahindra Docket Audit Tool - CV</h1>"
    "<p style='font-size:18px; font-style:italic; font-weight:700; color:inherit; margin-top:-12px; margin-left:4px; '>by Nishal Modi</p>"
    ,
    unsafe_allow_html=True
)

# --- Excel File Selection ---
selected_file_label = st.selectbox("📅 Select Excel File", file_labels, key="main_excel_select")
selected_filepath = os.path.join(DATA_DIR, file_map[selected_file_label])

# --- Load Data ---
data = pd.read_excel(selected_filepath, sheet_name=SHEET_NAME, header=HEADER_ROW)
data.drop(data.columns[0], axis=1, inplace=True)
data.columns = [str(col).strip().replace("\n", " ").replace("  ", " ") for col in data.columns]

# --- Variant Dropdown with Reset ---
current_variants = data["Variant"].dropna().drop_duplicates().tolist()
if "selected_variant" not in st.session_state:
    st.session_state.selected_variant = None

if st.session_state.selected_variant not in current_variants:
    st.session_state.selected_variant = current_variants[0] if current_variants else None

selected_variant = st.selectbox(
    "🎯 Select Vehicle Variant",
    current_variants,
    index=current_variants.index(st.session_state.selected_variant),
    key="variant_selectbox"
)
st.session_state.selected_variant = selected_variant

# --- Filter by Variant ---
filtered = data[data["Variant"] == selected_variant]
if filtered.empty:
    st.warning("⚠️ No data found for selected variant.")
    st.stop()
row = filtered.iloc[0]

# --- Currency Formatter ---
def format_indian_currency(value):
    try:
        if pd.isnull(value) or value == 0:
            return "₹0"
        value = float(value)
        is_negative = value < 0
        value = abs(value)
        s = f"{int(value)}"
        last_three = s[-3:]
        other = s[:-3]
        if other:
            other = re.sub(r'(\d)(?=(\d{2})+$)', r'\1,', other)
            formatted = f"{other},{last_three}"
        else:
            formatted = last_three
        result = f"₹{formatted}"
        return f"-{result}" if is_negative else result
    except:
        return "Invalid"

# --- Dynamic Cartel Group Extraction from Excel (Merged Cells) ---
def extract_cartel_groups(excel_path, sheet_name, header_row_idx):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    column_header_row = header_row_idx + 1
    group_header_row = column_header_row - 1

    headers = {
        col: ws.cell(row=column_header_row, column=col).value
        for col in range(1, ws.max_column + 1)
    }

    # Find cartel start column
    start_col = None
    for col_idx, val in headers.items():
        if val == "ON ROAD PRICE Without SMC Road Tax":
            start_col = col_idx + 1
            break

    if not start_col:
        return []

    end_col = ws.max_column
    cartel_groups = []

    # ✅ Sort merged cells left → right
    merged_ranges = sorted(
        ws.merged_cells.ranges,
        key=lambda r: r.bounds[0]
    )

    for merged in merged_ranges:
        min_col, min_row, max_col, max_row = merged.bounds

        if min_row != group_header_row:
            continue

        if max_col < start_col or min_col > end_col:
            continue

        group_name = ws.cell(row=group_header_row, column=min_col).value
        if not group_name:
            continue

        cols = []
        for c in range(max(min_col, start_col), min(max_col, end_col) + 1):
            header = headers.get(c)
            if header:
                cols.append(header)

        if cols:
            cartel_groups.append((group_name, cols))

    return cartel_groups


# --- Selected Variant Title ---
#st.markdown(f"<h2 style='margin-top: -8px; '> 🚚 {selected_variant}", unsafe_allow_html=True)

# --- Pricing Table ---
st.markdown("<h3 style='color:#e65100; margin-bottom: -8px;'>📝 Vehicle Pricing Details</h3>", unsafe_allow_html=True)

# Modify columns: Remove "MAXI CARE"
vehicle_cols = [
    "Ex-Showroom Price", "TCS", "Comprehensive + Zero Dep. Insurance",
    "R.T.O. Charges With Hypo.", "RSA (Road Side Assistance) For 1 Year",
    "SMC Road - Tax (If Applicable)", "Accessories",
    "ON ROAD PRICE With SMC Road Tax", "ON ROAD PRICE Without SMC Road Tax"
]

# Adjust ON ROAD PRICE values by subtracting MAXI CARE
adjusted_row = row.copy()
maxi_care_value = row.get("MAXI CARE", 0)
if pd.notnull(maxi_care_value):
    adjusted_row["ON ROAD PRICE With SMC Road Tax"] -= maxi_care_value
    adjusted_row["ON ROAD PRICE Without SMC Road Tax"] -= maxi_care_value

# Render pricing table
pricing_html = """
<style>
.vtable { border-collapse: collapse; width: 100%; font-weight: bold; font-size: 14px; }
.vtable th { background-color: #004080; color: white; padding: 4px 6px; text-align: right; }
.vtable td { background-color: #f0f4f8; padding: 4px 6px; text-align: right; color: black; }
.vtable td:first-child, .vtable th:first-child { text-align: left; }
.vtable, .vtable th, .vtable td { border: 1px solid #000; }
</style>
<table class='vtable'><tr><th>Description</th><th>Amount</th></tr>
"""
for col in vehicle_cols:
    pricing_html += f"<tr><td>{col}</td><td>{format_indian_currency(adjusted_row[col])}</td></tr>"
pricing_html += "</table>"
st.markdown(pricing_html, unsafe_allow_html=True)


# --- Cartel Offer (Single Table with Navy Group Headers) ---

cartel_groups = extract_cartel_groups(
    selected_filepath,
    SHEET_NAME,
    HEADER_ROW
)

st.markdown(
    "<h3 style='color:#e65100; margin-top:8px;'>🎁 Cartel Offer</h3>",
    unsafe_allow_html=True
)

if not cartel_groups:
    st.warning("⚠️ No cartel offer data found.")
else:
    cartel_html = """
    <style>
        table.ctable {
            border-collapse: collapse;
            width: 100%;
            font-weight: bold;
            font-size: 14px;
        }
        table.ctable th {
            background-color: #2e7d32;
            color: white;
            padding: 4px 6px;
            text-align: right;
            border: 1px solid #000;
        }
        table.ctable td {
            background-color: #e8f5e9;
            padding: 4px 6px;
            text-align: right;
            color: black;
            border: 1px solid #000;
        }
        table.ctable td:first-child,
        table.ctable th:first-child {
            text-align: left;
        }
        table.ctable tr.group-header td {
            background-color: transparent !important;
            color: #0b3c5d;  /* Navy Blue */
            font-weight: bold;
            text-align: left;
            border-top: 2px solid #000;
        }
    </style>

    <table class="ctable">
        <tr>
            <th>Description</th>
            <th>Offer</th>
        </tr>
    """

    for group_name, cols in cartel_groups:

        # Group Header
        cartel_html += f"""
        <tr class="group-header">
            <td colspan="2">{group_name}</td>
        </tr>
        """

        for col in cols:
            val = row.get(col)

            if isinstance(val, (int, float)):
                val = format_indian_currency(val)
            elif pd.isna(val):
                val = "₹0"
            else:
                val = str(val)

            cartel_html += f"""
            <tr>
                <td>{col}</td>
                <td>{val}</td>
            </tr>
            """

    cartel_html += "</table>"

    # ⚠️ THIS LINE IS CRITICAL
    st.markdown(cartel_html, unsafe_allow_html=True)


# --- Important Points Table ---
try:
    points_df = pd.read_excel(
        selected_filepath,
        sheet_name="Report",
        header=None,
        usecols="F:G",   # ✅ F = Sr. , G = Points
        skiprows=5,      # ✅ Skip first 5 rows (start from row 6)
        nrows=20         # ✅ Max till row 25
    ).dropna()

    # Rename columns
    points_df.columns = ["Sr.", "Points"]

    # Subtitle
    st.markdown(
        "<h3 style='color:#e65100; margin-top: -10px; margin-bottom: -8px;'>⭐ Important Points</h3>",
        unsafe_allow_html=True
    )

    # Build HTML table (use global styling)
    points_html = "<table class='iptable'><tr><th>Sr.</th><th>Points</th></tr>"
    for _, row in points_df.iterrows():
        points_html += f"<tr><td style='text-align:center'>{int(row['Sr.'])}</td><td>{row['Points']}</td></tr>"
    points_html += "</table>"

    st.markdown(points_html, unsafe_allow_html=True)

except Exception as e:
    st.warning(f"⚠️ Could not load Important Points: {e}")
